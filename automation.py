import os
import math
from openpyxl import load_workbook

SETTINGS = {'vlv': (['VE'], [''], ['']), 'vlv_on': (['B'], [''], ['_E1']), 'vlv_off': (['B'], [''], ['_E1']),
            'vlv_e_su': ([''], [''], ['_SU', 'SU']),
            'vlv_e_so': ([''], [''], ['_SO', 'SO']),
            'mot': (['PU', 'RW', 'HP', 'P', 'M'], [''], ['']),
            'mot_ra': ([''], [''], ['_E1']),
            'mot1R2G_L': ([''], [''], ['_L']),
            'mot1R2G_S': ([''], [''], ['_S']),
            'mot1R2G_L_ra': ([''], ['RA_'], ['_L']),
            'mot1R2G_S_ra': ([''], ['RA_'], ['_S']),
            'fu_ready': ([''], [''], ['_FU_BEREIT']),
            'fu_run': ([''], [''], ['_FU_M_LAEUFT']),
            'rel': ([''], [''], ['_N']),
            'ai': ([''], [''], ['_AE']),
            'ao': ([''], [''], ['_AA', '_SET']),
            'setpoint': ([''], [''], ['_AA', '_SET']),
            'pid': (['IC', 'IRC', 'ITC', 'ICR', 'CR'], [''], ['']),
            'visu_fc_num': 7000, 'ctrl_fc_num': 6000, 'idb_num': 7000, 'idb_num_ai': 4001, 'meas_fc_num': 4000,
            'idb_num_ao': 5001, 'fc_num_ao': 5000, 'idb_num_pid': 3501, 'idb_num_pid_ctrl': 3751, 'fc_num_pid': 3501,
            'start_analog': 1024, 'end_analog': 2200, 'start_digital1': 0, 'start_digital2': 4096, 'end_digital2': 4096,
            'tia_connection': 'HMI_Verbindung_1'}

SETTINGSCIP = {'vlv': (['VE'], [''], ['']), 'vlv_on': ([''], [''], ['_E']), 'vlv_off': ([''], [''], ['_XXX']),
            'vlv_e_su': ([''], [''], ['_SU', 'SU']),
            'vlv_e_so': ([''], [''], ['_SO', 'SO']),
            'mot': (['PU', 'RW', 'HP', 'P', 'M'], [''], ['']),
            'mot_ra': ([''], [''], ['_E']),
            'mot1R2G_L': ([''], [''], ['_L']),
            'mot1R2G_S': ([''], [''], ['_S']),
            'mot1R2G_L_ra': ([''], ['RA_'], ['_L']),
            'mot1R2G_S_ra': ([''], ['RA_'], ['_S']),
            'fu_ready': ([''], [''], ['_FU_BEREIT']),
            'fu_run': ([''], [''], ['_FU_M_LAEUFT']),
            'rel': ([''], [''], ['_N']),
            'ai': (['LT', 'TT', 'PIC', 'CT', 'FT', 'FIR', 'FI', 'PI'], [''], ['']),
            'ao': ([''], [''], ['']),
            'setpoint': (['VE'], [''], ['']),
            'pid': (['IC', 'IRC', 'ITC'], [''], ['']),
            'visu_fc_num': 7000, 'ctrl_fc_num': 6000, 'idb_num': 7000, 'idb_num_ai': 4001, 'meas_fc_num': 4000,
            'idb_num_ao': 5001, 'fc_num_ao': 5000, 'idb_num_pid': 3501, 'idb_num_pid_ctrl': 3751, 'fc_num_pid': 3501,
            'start_analog': 1024, 'end_analog': 2048, 'start_digital1': 0, 'start_digital2': 4096, 'end_digital2': 4096,
            'tia_connection': 'HMI_Verbindung_1'}

class IOAddress:

    def __init__(self, *args):  # address string // area, byte, bit
        self.str_add = ''
        self.area = ''
        self.byte = 0
        self.bit = None
        if len(args) == 1:
            self.init_str(*args)
        else:
            self.init_det(*args)

    def init_str(self, *args):
        self.str_add = args[0]
        self.area = self.str_add.split(' ')[0]
        num_add = self.str_add.split(' ')[-1].split('.')
        self.byte = int(num_add[0])
        if len(num_add) > 1:
            self.bit = int(num_add[1])

    def init_det(self, *args):
        self.area = args[0]
        self.byte = args[1]
        if len(args) > 2:
            self.bit = args[2]
            str_bit = '.' + str(self.bit)
        else:
            str_bit = ''
        spaces = (9 - len(str(self.byte)) - len(self.area)) * ' '
        str_byte = str(self.byte)
        self.str_add = self.area + spaces + str_byte + str_bit

    def __str__(self):
        return self.str_add

    def is_output(self):
            return self.area in ['A', 'Q']

    def is_input(self):
        return self.area in ['E', 'I']

    def is_analog_input(self):
        return self.area in ['PEW', 'EW', 'PIW', 'IW']

    def is_analog_output(self):
        return self.area in ['PAW', 'AW', 'PQW', 'QW']

    def get_id(self):
        if self.is_analog_input() or self.is_analog_output():
            return (self.byte-SETTINGS['start_analog'])//2
        else:
            if self.byte >= SETTINGS['start_digital2']:
                return (self.byte-SETTINGS['start_digital2'])*8+SETTINGS['start_digital2']+self.bit
            else:
                return (self.byte-SETTINGS['start_digital1'])*8+self.bit

    def get_var_type(self):
        var_types ={'A': 'BOOL', 'Q': 'BOOL', 'E': 'BOOL', 'I': 'BOOL', 'M': 'BOOL', 'MB': 'BYTE',
                    'AW': 'WORD', 'QW': 'WORD', 'EW': 'WORD', 'IW': 'WORD', 'PAW': 'WORD', 'PQW': 'WORD',
                    'PEW': 'WORD', 'PIW': 'WORD', 'MW': 'WORD', 'MD': 'DWORD', 'FC': self.str_add,
                    'FB': self.str_add}
        try:
            return var_types[self.area]
        except KeyError:
            return 'UNKNOWN'


class Symbol:
    def __init__(self, symbol, comment='', var_type=None, *args):  # address string // area, byte, bit
        self.symbol = symbol
        self.address = IOAddress(*args)
        if var_type:
            self.var_type = var_type
        else:
            self.var_type = self.address.get_var_type()
        if comment:
            self.comment = comment[:80]
        else:
            self.comment = ""

    def __str__(self):
        return self.symbol

    def to_file(self):
        return self.symbol+'\t'+str(self.address)+'\t'+str(self.var_type)+'\t'+self.comment

    def to_listbox(self):
        return str(self.symbol)

    def to_lib_comment(self):
        if self.is_output():
            str_comm = self.comment if self.comment else "C{0:04}.{1:1}".format(self.address.byte, self.address.bit)
            return str(self.address.get_id())+'\t'+str_comm
        else:
            str_comm = self.comment if self.comment else "C_{0}{1}".format(self.address.area, self.address.byte)
            return str(self.address.get_id())+'\t'+str_comm

    def to_lib_address(self):
        if self.is_output():
            str_adr = "{0}{1:03}.{2:1}".format(self.address.area, self.address.byte, self.address.bit)
            return str(self.address.get_id())+'\t'+str_adr
        else:
            str_adr = "{0}{1}".format(self.address.area, self.address.byte)
            return str(self.address.get_id())+'\t'+str_adr

    def is_output(self):
        return self.address.is_output()

    def is_input(self):
        return self.address.is_input()

    def is_analog_input(self):
        return self.address.is_analog_input() and self.symbol_test(SETTINGS['ai'])

    def is_pid(self):
        return self.symbol_test(SETTINGS['pid'])

    def is_analog_output(self):
        return self.address.is_analog_output() and self.symbol_test(SETTINGS['ao'])

    def is_setpoint(self):
        return self.is_analog_output() and self.symbol_test(SETTINGS['setpoint'])

    def symbol_test(self, setting):
        cont, beg, end = setting
        beg_ok = True
        end_ok = True
        cont_ok = True
        if beg[0] != "":
            beg_ok = False
            for b in beg:
                if b and self.symbol[:len(b)] == b:
                    beg_ok=True
        if end[0] != "":
            end_ok = False
            for e in end:
                if e and self.symbol[-len(e):] == e:
                    end_ok=True
        if cont[0] != "":
            cont_ok = False
            for c in cont:
                if c and c in self.symbol:
                    if self.symbol.find(c) < len(self.symbol) - len(c):
                        if self.symbol[self.symbol.find(c)+len(c)] in [str(n) for n in range(0, 10)]:
                            cont_ok=True
        # print("symbol test: {0}, setting: {1}, result: beg {2}, end {3}, cont {4}".format(self.symbol, setting, beg_ok, end_ok, cont_ok))
        return beg_ok and end_ok and cont_ok and (beg or end or cont)

    def is_relay(self):
        return self.symbol_test(SETTINGS['rel'])

    def is_valve(self):
        return self.symbol_test(SETTINGS['vlv'])

    def is_valve_e(self):
            return self.is_valve_e_su() or self.is_valve_e_so()

    def is_valve_e_out(self):
        return self.is_valve and not (self.is_valve_e_su() or self.is_valve_e_so())

    def is_valve_e_su(self):
        return self.is_valve() and self.symbol_test(SETTINGS['vlv_e_su'])

    def is_valve_e_so(self):
        return self.is_valve() and self.symbol_test(SETTINGS['vlv_e_so'])

    def is_motor(self):
        return self.symbol_test(SETTINGS['mot'])

    def is_motor1r2g(self):
        return self.is_motor1r2g_l() or self.is_motor1r2g_s()

    def is_motor1r2g_l(self):
        return self.is_motor() and self.symbol_test(SETTINGS['mot1R2G_L'])

    def is_motor1r2g_s(self):
        return self.is_motor() and self.symbol_test(SETTINGS['mot1R2G_S'])

    def is_valve_input(self):
        return self.is_valve_on() or self.is_valve_off()

    def is_valve_on(self):
        return self.symbol_test(SETTINGS['vlv_on'])

    def is_valve_off(self):
        return self.symbol_test(SETTINGS['vlv_off'])

    def is_mot_input(self):
        return self.is_mot_on()

    def is_mot_on(self):
        return self.symbol_test(SETTINGS['mot_ra'])

    def is_mot_1r2g_input(self):
        return self.is_mot_l_on() or self.is_mot_s_on()

    def is_mot_l_on(self):
        return self.symbol_test(SETTINGS['mot1R2G_L_ra'])

    def is_mot_s_on(self):
        return self.symbol_test(SETTINGS['mot1R2G_S_ra'])

    def is_fu(self):
        return self.is_fu_ready() or self.is_fu_run()

    def is_fu_ready(self):
        return self.symbol_test(SETTINGS['fu_ready'])

    def is_fu_run(self):
        return self.symbol_test(SETTINGS['fu_run'])

    def make_name(self):
        if self.is_output():
            if self.is_motor1r2g_l():
                return self.symbol_to_name(SETTINGS['mot1R2G_L'])
            if self.is_motor1r2g_s():
                return self.symbol_to_name(SETTINGS['mot1R2G_S'])
            if self.is_motor():
                return self.symbol_to_name(SETTINGS['mot'])
            if self.is_valve_e_so():
                return self.symbol_to_name(SETTINGS['vlv_e_so'])
            if self.is_valve_e_su():
                return self.symbol_to_name(SETTINGS['vlv_e_su'])
            if self.is_valve():
                return self.symbol_to_name(SETTINGS['vlv'])
            if self.is_relay():
                return self.symbol.replace('', '_A1')
            return self.symbol_to_name(([''], [''], ['']))

        elif self.is_analog_input():
            return self.symbol.replace('_AE', '')
        elif self.is_analog_output():
            return self.symbol.replace('_AA', '_SP').replace('_SET', '_SP')
        elif self.is_input():
            if self.is_fu_ready():
                return self.symbol_to_name(SETTINGS['fu_ready'])
            elif self.is_fu_run():
                return self.symbol_to_name(SETTINGS['fu_run'])
            elif self.is_valve_on():
                return self.symbol_to_name(SETTINGS['vlv_on']).replace("B","VE")
            elif self.is_valve_off():
                return self.symbol_to_name(SETTINGS['vlv_off']).replace("B","VE")
            elif self.is_mot_on():
                return self.symbol_to_name(SETTINGS['mot_ra'])
        else:
            return ''

    def symbol_to_name(self, setting):
        cont, beg, end = setting
        sym = self.symbol
        if self.is_output():
            sym = sym.replace('_A1', '')
        elif self.is_analog_input():
            sym = sym.replace('_AE', '')
        elif self.is_analog_output():
            sym = sym.symbol.replace('_AA', '')
        for b in beg:
            if b:
                sym = sym.replace(b, '')
        for e in end:
            if e:
                sym = sym.replace(e, '')
        return sym

    def make_operand(self, var_type='BOOL', init='FALSE'):
        symbol = self.make_name()
        if symbol:
            return Operand(symbol, var_type, self.comment, init)
        return None

    def is_component(self):
        return self.is_output()

    def find_other_output(self, symbols, setting):
        cont, beg, end = setting
        outputs = [s for s in symbols if s.is_output()]
        output_symbols = [s.symbol for s in outputs]
        for b in beg:
            for e in end:
                o_name = b+self.make_name()+e
                if o_name in output_symbols:
                    return outputs[output_symbols.index(o_name)]

    def make_component(self, components, names, symbols=None):
        if self.is_output():
            if self.is_relay():
                components.append(Relay(name=self.make_name(), comment=self.comment, io_out=self))

            elif self.is_valve_e_so():
                if self.make_name() in names:
                    idx = names.index(self.make_name())
                    if components[idx].__class__.__name__ == "Valve":
                        io_out = components[idx].io_out
                        io_on = components[idx].io_on
                        io_off = components[idx].io_off
                        components[idx] = ValveE(name=self.make_name(), comment=self.comment, io_out=io_out,
                                                 io_on=io_on, io_off=io_off, io_out_o=self)
                    else:
                        components[idx].io_out_o = self
                else:
                    io_out = self.find_other_output(symbols, SETTINGS['vlv'])
                    components.append(ValveE(name=self.make_name(), comment=self.comment, io_out=io_out, io_out_o=self))

            elif self.is_valve_e_su():
                if self.make_name() in names:
                    idx = names.index(self.make_name())
                    if components[idx].__class__.__name__ == "Valve":
                        io_out = components[idx].io_out
                        io_on = components[idx].io_on
                        io_off = components[idx].io_off
                        components[idx] = ValveE(name=self.make_name(), comment=self.comment, io_out=io_out,
                                                 io_on=io_on, io_off=io_off, io_out_u=self)
                    else:
                        components[idx].io_out_u = self
                else:
                    io_out = self.find_other_output(symbols, SETTINGS['vlv'])
                    components.append(ValveE(name=self.make_name(), comment=self.comment, io_out=io_out, io_out_u=self))

            elif self.is_valve():
                if not self.make_name() in names:
                    components.append(Valve(name=self.make_name(), comment=self.comment, io_out=self))

            elif self.is_motor1r2g_l():
                if not self.make_name() in names:
                    io_out_s = self.find_other_output(symbols, SETTINGS['mot1R2G_S'])
                    components.append(Motor1R2G(name=self.make_name(), comment=self.comment, io_out_l=self, io_out_s=io_out_s,
                                     symbols=symbols))
                else:
                    idx = names.index(self.make_name())
                    components[idx].io_out_l = self

            elif self.is_motor1r2g_s():
                if not self.make_name() in names:
                    return Motor1R2G(name=self.make_name(), comment=self.comment, io_out_s=self, symbols=symbols)

            elif self.is_motor():
                components.append(Motor(name=self.make_name(), comment=self.comment, io_out=self, symbols=symbols))
            else:
                components.append(Placeholder(name=self.make_name(), comment=self.comment, io_out=self))

    def update_component(self, components, names):
        if self.is_input():
            # print(self)
            # print(self.make_name())
            if self.make_name() in names:
                idx = names.index(self.make_name())
                if self.is_valve_on() and components[idx].__class__.__name__ in ["Valve", "ValveE"]:
                    components[idx].io_on = self
                elif self.is_valve_off() and components[idx].__class__.__name__ in ["Valve", "ValveE"]:
                    components[idx].io_off = self
                elif self.is_mot_on() and components[idx].__class__.__name__ in ["Motor"]:
                    components[idx].io_on = self
                elif self.is_mot_l_on() and components[idx].__class__.__name__ in ["Motor_1R2G"]:
                    components[idx].io_on_l = self
                elif self.is_mot_s_on() and components[idx].__class__.__name__ in ["Motor_1R2G"]:
                    components[idx].io_on_s = self
                elif self.is_fu():
                    if components[idx].__class__.__name__ == "Motor":
                        io_out = components[idx].io_out
                        if self.is_fu_ready:
                            components[idx] = MotorFU(name=self.make_name(), comment=self.comment, io_out=io_out,
                                                      io_ready=self)
                        elif self.is_fu_run:
                            components[idx] = MotorFU(name=self.make_name(), comment=self.comment, io_out=io_out,
                                                      io_run=self)
                    elif self.is_fu_ready():
                        components[idx].io_ready = self
                    elif self.is_fu_run():
                        components[idx].io_run = self

    def make_measurement(self):
        if self.is_analog_input():
            return Measurement(name=self.make_name(), comment=self.comment, io_in=self, lib_id=self.address.get_id())

    def make_analog_output(self):
        if self.is_analog_output():
            if self.is_setpoint():
                return Setpoint(name=self.make_name(), comment=self.comment, io_out=self, lib_id=self.address.get_id())
            return AoPlaceholder(name=self.make_name(), comment=self.comment, io_out=self, lib_id=self.address.get_id())


class Operand:
    def __init__(self, symbol, var_type='BOOL', comment='', init='FALSE'):
        self.symbol = symbol.replace(' ', '').replace('.', '_').replace('-', '_').replace('+', '_')
        if self.symbol[0].isdigit():
            self.symbol = '_' + self.symbol
        if self.symbol[-1] == '_':
            self.symbol = self.symbol[:-2]
        if self.symbol[0:3] == 'AI_':
            self.symbol = self.symbol[3:]
        if self.symbol[0:3] == 'AO_':
            self.symbol = self.symbol[3:]
        self.var_type = var_type
        self.comment = comment
        self.init = init

    def __str__(self):
        str_comment = ("\t//" + self.comment) if self.comment else ''
        return self.symbol + " : " + self.var_type + ";" + str_comment

    def str_init(self, struct_name=None):
        symbol_name = self.symbol
        if struct_name:
            symbol_name = struct_name+'.'+symbol_name
        return symbol_name + " := " + self.init + ";"


class Struct:
    def __init__(self, name, comment='', operands=None):
        self.name = name
        self.operands = operands
        self.comment = comment

    def __str__(self):
        try:
            operands_string = '\n'.join(["\t"+str(o) for o in self.operands])
        except TypeError:
            operands_string = "\t"+str(self.operands)
        return '{} : STRUCT //{}\n'.format(self.name, self.comment) + operands_string + '\nEND_STRUCT;'

    def str_init(self, struct_name=None):
        symbol_name = self.name
        if struct_name:
            symbol_name = struct_name+'.'+symbol_name
        try:
            return '\n'.join(["\t"+o.str_init(symbol_name) for o in self.operands])
        except TypeError:
            return "\t"+self.operands.str_init(symbol_name)


class Array:
    def __init__(self, name, start, end, var_type, init, comment=''):
        self.name = name
        self.start = start
        self.end = end
        self.var_type = var_type
        self.init = init
        self.comment = comment

    def __str__(self):
        return '{} : ARRAY  [{} .. {} ] OF {} ; //{}\n'.format(self.name, self.start, self.end, self.var_type,
                                                               self.comment)

    def str_init(self, struct_name=None):
        array_inits = []
        for num in range(self.start, self.end+1):
            symbol_name = '{}[{}] '.format(self.name, num)
            if struct_name:
                symbol_name = struct_name+'.'+symbol_name
            array_inits.append(symbol_name + " := " + self.init + ";")
        return '\n'.join(["\t"+init for init in array_inits])


class Component:
    FB = None

    def __init__(self, name, comment):
        self.name = name
        self.comment = comment
        self.lib_id = 0

    def __str__(self):
        return self.name

    def to_listbox(self):
        return self.__class__.__name__+':  '+self.name

    def find_input(self, symbols, setting):
        cont, beg, end = setting
        inputs = [s for s in symbols if s.is_input()]
        input_symbols = [s.symbol for s in inputs]
        for b in beg:
            for e in end:
                e_name = b+self.name+e
                if e_name in input_symbols:
                    return inputs[input_symbols.index(e_name)]

        # if cont:
        #     for c in cont:
        #         b_name = self.name.replace(c, 'B')
        #         if b_name in input_symbols:
        #             return inputs[input_symbols.index(b_name)]

        return None

    def find_output(self, symbols, setting):
        cont, beg, end = setting
        outputs = [s for s in symbols if s.is_output()]
        output_symbols = [s.symbol for s in outputs]
        for b in beg:
            for e in end:
                o_name = "A_"+b+self.name+e
                if o_name in output_symbols:
                    return outputs[output_symbols.index(o_name)]

    def make_operand(self):
        return Operand(self.name, 'BOOL', self.comment)

    def make_ctrl_network(self):
        network_code = """
        CLR;
        =     "DB_AUSGABE".{0};

        """.format(self.make_operand().symbol)
        return FCNetwork(title=self.name, code=network_code, byte=self.io_out.address.byte)

    def make_tiavariables(self):
        return ""

    def make_tiaalarms(self):
        return ""

    def make_tiaarchive(self):
        return ""

    def __str__(self):
        return str(self.make_network())

class Valve(Component):
    FB = 560

    def __init__(self, name, comment, io_out, io_on=None, io_off=None, no=False):
        super(Valve, self).__init__(name, comment)
        self.io_out = io_out
        self.io_on = io_on
        self.io_off = io_off
        self.no = no

        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = SETTINGS['idb_num']+self.io_out.address.get_id()
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def find_input_on(self, symbols, setting):
        self.io_on = super(Valve, self).find_input(symbols, setting)

    def find_input_off(self, symbols, setting):
        self.io_off = super(Valve, self).find_input(symbols, setting)

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            str_io_off = 'RM_N_ANST       := "{0}",'.format(self.io_off) if self.io_off \
                else 'RM_N_ANST_FRG   := FALSE,'
            str_io_on = 'RM_ANST         := "{0}",'.format(self.io_on) if self.io_on \
                else 'RM_ANST_FRG     := FALSE,'

            network_code = """        CALL  "VENTIL_8P" , "{0}" (
           IMP_ZEITBASIS   :="AUSG_IMP_ZEITBASIS",
           STOER_AUS       :="STOER_AUS",
           QUITT_EX        :="QUITT_EX",
           ANST_EX_AUTO    :="DB_AUSGABE".{1},
           {2}
           {3}
           FRG_TSE         :=FALSE,
           FRG_TSA         :=FALSE,
           ZEITBASIS       :="DB_SYSTEM".AUSG_IMP_ZEITBASIS,
           BIBLIOTHEK_ID   :=L#{5},
           STOER           :="DB_AUSG_STOER".{1},
           STOER_SP        :="DB_AUSG_STOER_SP".{1},
           BA              :="DB_AUSG_BA".{1},
           AUSGANG         :="{4}",
           ENDLAGE_NA      :="DB_AUSG_ENDL_NA".{1},
           ENDLAGE_ANG     :="DB_AUSG_ENDL_ANG".{1},
           Para_Werkseinst :=#para_werks);
    // 
          U     #para_werks;
          O     #para_werks_sammel;
          SPBN  M{6:02}z;
    // 
          CALL  "VENTIL_WE" (
           IDB            :={7},
           Werkseinst     :=TRUE,
           SZ_ANST_EIN    :=1.000000e+001,
           SZ_ANST_AUS    :=1.000000e+001,
           SZ_TSE         :=0.000000e+000,
           SZ_TSA         :=0.000000e+000,
           GW_LZZ         :=L#0,
           GW_SSZ         :=L#0,
           FRG_E_VERZ     :=FALSE,
           FRG_A_VERZ     :=FALSE,
           STOER_AUS      :=FALSE,
           FRG_H_INT      :=TRUE,
           FRG_H_EXT      :=FALSE,
           ANTR_ST_AUS    :=FALSE,
           RM_NA_UEB_AKT  :=FALSE,
           RM_ANST_UEB_AKT:=TRUE,
           BED_O_PW       :=FALSE);
        // 
        M{6:02}z: NOP   0;""".format(self.name, self.make_operand().symbol, str_io_off, str_io_on, str(self.io_out),
                       str(self.lib_id), self.network_no, self.idb_num)
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def make_tiavariables(self):

            variables = """{0}.BIBLIOTHEK_ID	{1}	{2}	{0}.BIBLIOTHEK_ID	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD8	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.MSG_EVID	{1}	{2}	{0}.MSG_EVID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD12	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CHARGE	{1}	{2}	{0}.CHARGE	String	String	18	Binary	Absolute access	%DB{3}.DBX16.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BEREICH	{1}	{2}	{0}.BEREICH	String	String	18	Binary	Absolute access	%DB{3}.DBX36.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.RECHNERNAME	{1}	{2}	{0}.RECHNERNAME	String	String	16	Binary	Absolute access	%DB{3}.DBX56.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CURRENT_USER	{1}	{2}	{0}.CURRENT_USER	String	String	20	Binary	Absolute access	%DB{3}.DBX74.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.obyANST	{1}	{2}	{0}.obyANST	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB102	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwRM_DT	{1}	{2}	{0}.odwRM_DT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD104	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwIN	{1}	{2}	{0}.odwIN	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD108	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwOUT	{1}	{2}	{0}.odwOUT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD112	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwEV_ID	{1}	{2}	{0}.iodwEV_ID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD118	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwDT	{1}	{2}	{0}.iodwDT	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD122	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_ANST_EIN	{1}	{2}	{0}.iorSZ_ANST_EIN	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD126	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_ANST_AUS	{1}	{2}	{0}.iorSZ_ANST_AUS	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD130	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_ANST	{1}	{2}	{0}.iorIZ_ANST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD134	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE	{1}	{2}	{0}.iorSZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD138	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSE	{1}	{2}	{0}.iorIZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD142	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSA	{1}	{2}	{0}.iorSZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD146	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSA	{1}	{2}	{0}.iorIZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD150	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiLZZ	{1}	{2}	{0}.iodiLZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD154	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiGW_LZZ	{1}	{2}	{0}.iodiGW_LZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD158	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiSSZ	{1}	{2}	{0}.iodiSSZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD162	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiGW_SSZ	{1}	{2}	{0}.iodiGW_SSZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD166	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owAlarm	{1}	{2}	{0}.owAlarm	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW116	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iowQuitt	{1}	{2}	{0}.iowQuitt	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW170	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False"""\
                .format(self.name, self.__class__.__name__, SETTINGS['tia_connection'], self.idb_num)
            return variables

    def make_tiaalarms(self):

        alarms = '''{0}_Laufzeit	Fehler laufzeit {0}		Errors	"""{0}.owAlarm"""	0	On rising edge	"""{0}.iowQuitt"""	0	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Extern	Fehler extern {0}		Errors	"""{0}.owAlarm"""	1	On rising edge	"""{0}.iowQuitt"""	1	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Wartung	Wartung {0}		Warnings	"""{0}.owAlarm"""	2	On rising edge	"""{0}.iowQuitt"""	2	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Betriebsart	Betriebsart {0}		Hand	"""{0}.owAlarm"""	3	On rising edge	"""{0}.iowQuitt"""	3	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Ansteuerung	Ansteuerung {0}		Ansteurung	"""{0}.owAlarm"""	4	On rising edge	"""{0}.iowQuitt"""	4	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0'''\
            .format(self.name)
        return alarms

    def __str__(self):
        return str(self.make_network())


class ValveE(Component):
    FB = 561

    def __init__(self, name, comment, io_out, io_on=None, io_off=None, io_out_u=None,
                 io_out_o=None, no=False):
        super(ValveE, self).__init__(name, comment)
        self.io_out = io_out
        self.io_on = io_on
        self.io_out_u = io_out_u
        self.io_out_o = io_out_o
        self.io_off = io_off
        self.no = no
        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = SETTINGS['idb_num']+self.io_out.address.get_id()
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            str_io_off = 'RM_N_ANST       := "{0}",'.format(self.io_off) if self.io_off \
                else 'RM_N_ANST_FRG   := FALSE,'
            str_io_on = 'RM_ANST         := "{0}",'.format(self.io_on) if self.io_on \
                else 'RM_ANST_FRG     := FALSE,'
            str_io_out_o = 'ANST_LECKG_O    := "{0}",'.format(self.io_out_o) if self.io_out_o \
                else ''
            str_io_out_u = 'ANST_LECKG_U     := "{0}",'.format(self.io_out_u) if self.io_out_u \
                else ''

            network_code = """        CALL  "VENTIL_E_8P" , "{0}" (
           IMP_ZEITBASIS   :="AUSG_IMP_ZEITBASIS",
           STOER_AUS       :="STOER_AUS",
           QUITT_EX        :="QUITT_EX",
           ANST_EX_AUTO    :="DB_AUSGABE".{1},
           {2}
           {3}
           {4}
           {5}
           FRG_TSE         :=FALSE,
           FRG_TSA         :=FALSE,
           ZEITBASIS       :="DB_SYSTEM".AUSG_IMP_ZEITBASIS,
           BIBLIOTHEK_ID   :=L#{7},
           STOER           :="DB_AUSG_STOER".{1},
           STOER_SP        :="DB_AUSG_STOER_SP".{1},
           BA              :="DB_AUSG_BA".{1},
           AUSGANG         :="{6}",
           ENDLAGE_NA      :="DB_AUSG_ENDL_NA".{1},
           ENDLAGE_ANG     :="DB_AUSG_ENDL_ANG".{1},
           Para_Werkseinst :=#para_werks);
    // 
          U     #para_werks;
          O     #para_werks_sammel;
          SPBN  M{8:02}z;
    // 
          CALL  "VENTIL_E_WE" (
           IDB            :={9},
           Werkseinst     :=TRUE,
           SZ_ANST_EIN    :=1.000000e+001,
           SZ_ANST_AUS    :=1.000000e+001,
           SZ_TSE         :=0.000000e+000,
           SZ_TSA         :=0.000000e+000,
           GW_LZZ         :=L#0,
           GW_SSZ         :=L#0,
           FRG_E_VERZ     :=FALSE,
           FRG_A_VERZ     :=FALSE,
           STOER_AUS      :=FALSE,
           FRG_H_INT      :=TRUE,
           FRG_H_EXT      :=FALSE,
           ANTR_ST_AUS    :=FALSE,
           FRG_LECKG_U    :=TRUE,
           FRG_LECKG_O    :=TRUE,
           RM_NA_UEB_AKT  :=FALSE,
           RM_ANST_UEB_AKT:=TRUE,
           LECKG_U_RM_EIN :=FALSE,
           LECKG_O_RM_EIN :=FALSE,
           BED_O_PW       :=FALSE);
        // 
        M{8:02}z: NOP   0;""".format(self.name, self.make_operand().symbol, str_io_out_u, str_io_out_o, str_io_off,
                                     str_io_on, str(self.io_out), str(self.lib_id), self.network_no, self.idb_num)
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def __str__(self):
        return str(self.make_network())


class Motor(Component):
    FB = 550

    def __init__(self, name, comment, io_out, symbols=None, io_on=None):
        super(Motor, self).__init__(name, comment)
        self.io_out = io_out  # Symbol
        self.io_on = io_on  # Symbol
        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = SETTINGS['idb_num']+self.io_out.address.get_id()
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            str_io_on = '\n            RM_ANST_EIN              := "{0}",'.format(self.io_on) if self.io_on \
                else '\n            RM_ANST_EIN_FRG          := FALSE,'

            network_code = """        CALL "MOTOR_1R1G_8P" , "{0}" (
            IMP_ZEITBASIS            := "AUSG_IMP_ZEITBASIS",
            FRG                      := TRUE,
            STOER_AUS                := "STOER_AUS",
            QUITT_EX                 := "QUITT_EX",
            ANST_EX_AUTO             := "DB_AUSGABE".{1},{2}
            FRG_TSE                  := TRUE,
            FRG_TSA                  := TRUE,
            MSS                      := FALSE,
            REPS                     := TRUE,
            MANNL                    := TRUE,
            ZEITBASIS                := "DB_SYSTEM".AUSG_IMP_ZEITBASIS,
            BIBLIOTHEK_ID            := L#{4},
            STOER                    := "DB_AUSG_STOER".{1},
            STOER_SP                 := "DB_AUSG_STOER_SP".{1},
            BA                       := "DB_AUSG_BA".{1},
            AUSGANG                  := "{3}",
            RM_EIN                   := "DB_AUSG_ENDL_ANG".{1},
            Para_Werkseinst          := #para_werks);
// 
      U     #para_werks; 
      O     #para_werks_sammel; 
      SPBN  M{5:02}z;
// 
      CALL "MOTOR_1R1G_WE" (
           IDB                      := {6},
           Werkseinst               := TRUE,
           SZ_ANST_EIN              := 1.000000e+001,
           SZ_ANST_AUS              := 1.000000e+001,
           SZ_TSE                   := 0.000000e+000,
           SZ_TSA                   := 0.000000e+000,
           GW_LZZ                   := L#0,
           GW_SSZ                   := L#0,
           FRG_E_VERZ               := FALSE,
           FRG_A_VERZ               := FALSE,
           STOER_AUS                := FALSE,
           FRG_H_INT                := TRUE,
           FRG_H_EXT                := FALSE,
           ANTR_ST_AUS              := TRUE,
           RM_UEB_AKT               := TRUE,
           BED_O_PW                 := FALSE);
// 
M{5:02}z: NOP   0;""".format(self.name, self.make_operand().symbol, str_io_on, str(self.io_out), str(self.lib_id),
                       self.network_no, self.idb_num)
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def make_tiavariables(self):
        
        variables = """{0}.CHARGE	{1}	{2}	{0}.CHARGE	String	String	18	Binary	Absolute access	%DB{3}.DBX16.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BEREICH	{1}	{2}	{0}.BEREICH	String	String	18	Binary	Absolute access	%DB{3}.DBX36.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.RECHNERNAME	{1}	{2}	{0}.RECHNERNAME	String	String	16	Binary	Absolute access	%DB{3}.DBX56.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CURRENT_USER	{1}	{2}	{0}.CURRENT_USER	String	String	20	Binary	Absolute access	%DB{3}.DBX74.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.obyANST	{1}	{2}	{0}.obyANST	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB102	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwRM_DT	{1}	{2}	{0}.odwRM_DT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD104	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwIN	{1}	{2}	{0}.odwIN	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD108	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwOUT	{1}	{2}	{0}.odwOUT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD112	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwEV_ID	{1}	{2}	{0}.iodwEV_ID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD118	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwDT	{1}	{2}	{0}.iodwDT	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD122	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_ANST_EIN	{1}	{2}	{0}.iorSZ_ANST_EIN	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD126	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_ANST_AUS	{1}	{2}	{0}.iorSZ_ANST_AUS	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD130	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_ANST	{1}	{2}	{0}.iorIZ_ANST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD134	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE	{1}	{2}	{0}.iorSZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD138	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSE	{1}	{2}	{0}.iorIZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD142	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSA	{1}	{2}	{0}.iorSZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD146	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSA	{1}	{2}	{0}.iorIZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD150	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiLZZ	{1}	{2}	{0}.iodiLZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD154	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiGW_LZZ	{1}	{2}	{0}.iodiGW_LZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD158	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiSSZ	{1}	{2}	{0}.iodiSSZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD162	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiGW_SSZ	{1}	{2}	{0}.iodiGW_SSZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD166	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owAlarm	{1}	{2}	{0}.owAlarm	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW116	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iowQuitt	{1}	{2}	{0}.iowQuitt	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW170	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False"""\
            .format(self.name, self.__class__.__name__, SETTINGS['tia_connection'], self.idb_num)
        return variables

    def make_tiaalarms(self):
        alarms = '''{0}_Laufzeit	Fehler laufzeit {0}		Errors	"""{0}.owAlarm"""	0	On rising edge	"""{0}.iowQuitt"""	0	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Extern	Fehler extern {0}		Errors	"""{0}.owAlarm"""	1	On rising edge	"""{0}.iowQuitt"""	1	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Wartung	Wartung {0}		Warnings	"""{0}.owAlarm"""	2	On rising edge	"""{0}.iowQuitt"""	2	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Betriebsart	Betriebsart {0}		Hand	"""{0}.owAlarm"""	3	On rising edge	"""{0}.iowQuitt"""	3	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Ansteuerung	Ansteuerung {0}		Ansteurung	"""{0}.owAlarm"""	4	On rising edge	"""{0}.iowQuitt"""	4	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0''' \
            .format(self.name)
        return alarms

    def __str__(self):
        return str(self.make_network())


class Motor1R2G(Component):
    FB = 552

    def __init__(self, name, comment, io_out_s, symbols=None, io_out_l=None, io_on_l=None, io_on_s=None):
        super(Motor1R2G, self).__init__(name, comment)
        self.io_out_l = io_out_l  # Symbol
        self.io_out_s = io_out_s  # Symbol
        self.io_on_l = io_on_l  # Symbol
        self.io_on_s = io_on_s  # Symbol
        self.lib_id = self.io_out_s.address.get_id()
        self.network_no = (self.io_out_s.address.byte % 4) * 8 + self.io_out_s.address.bit + 2
        self.idb_num = SETTINGS['idb_num']+self.io_out_s.address.get_id()
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out_s.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out_s.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out_l.address.get_id(), "operand": Operand(self.name+"_L", 'BOOL', self.comment)},
                {"id": self.io_out_s.address.get_id(), "operand": Operand(self.name+"_S", 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out_l and self.io_out_s:
            str_io_on = '\n            RM_ANST_L_EIN              := "{0}",' \
                        '\n            RM_ANST_S_EIN              := "{1}",'.format(self.io_on_l, self.io_on_s) \
                if self.io_on_l and self.io_on_s else '\n            RM_ANST_EIN_FRG          := FALSE,'

            network_code = """        CALL "MOTOR_1R2G_8P" , "{0}" (
            IMP_ZEITBASIS            := "AUSG_IMP_ZEITBASIS",
            FRG                      := TRUE,
            STOER_AUS                := "STOER_AUS",
            QUITT_EX                 := "QUITT_EX",
            ANST_EX_AUTO_L           := "DB_AUSGABE".{0}_l,
            ANST_EX_AUTO_S           := "DB_AUSGABE".{0}_S,{2}
            FRG_TSE                  := TRUE,
            FRG_TSA                  := TRUE,
            MSS                      := FALSE,
            REPS                     := TRUE,
            MANNL                    := TRUE,
            ZEITBASIS                := "DB_SYSTEM".AUSG_IMP_ZEITBASIS,
            BIBLIOTHEK_ID            := L#{5},
            STOER                    := "DB_AUSG_STOER".{0},
            STOER_SP                 := "DB_AUSG_STOER_SP".{0},
            BA                       := "DB_AUSG_BA".{0},
            AUSGANG_L                := "{3}",
            AUSGANG_S                := "{4}",
            RM_EIN                   := "DB_AUSG_ENDL_ANG".{1},
            Para_Werkseinst          := #para_werks);
               
// 
      U     #para_werks; 
      O     #para_werks_sammel; 
      SPBN  M{6:02}z;
// 
      CALL "MOTOR_1R2G_WE" (
           IDB                      := {7},
           Werkseinst               := TRUE,
           SZ_ANST_EIN              := 1.000000e+001,
           SZ_ANST_AUS              := 1.000000e+001,
           SZ_TSE                   := 0.000000e+000,
           SZ_TSA                   := 0.000000e+000,
           SZ_L_S                   := 0.000000e+000,
           SZ_S_L                   := 0.000000e+000,
           GW_LZZ                   := L#0,
           GW_SSZ                   := L#0,
           FRG_E_VERZ               := FALSE,
           FRG_A_VERZ               := FALSE,
           STOER_AUS                := FALSE,
           FRG_H_L_INT              := TRUE,
           FRG_H_S_INT              := TRUE,
           FRG_H_L_EXT              := FALSE,
           FRG_H_S_EXT              := FALSE,
           ANTR_ST_AUS              := TRUE,
           RM_UEB_AKT               := TRUE,
           BED_O_PW                 := FALSE);

// 
M{6:02}z: NOP   0;""".format(self.name, self.make_operand().symbol, str_io_on, str(self.io_out_l), str(self.io_out_s),
                       self.lib_id, self.network_no, self.idb_num)
            network_title = "A{0}.{1} / A{2}.{3} / {4}".format(str(self.io_out_l.address.byte),
                                                               str(self.io_out_l.address.bit),
                                                               str(self.io_out_s.address.byte),
                                                               str(self.io_out_s.address.bit),
                                                    self.name)

            return FCNetwork(title=network_title, code=network_code, byte=self.io_out_s.address.byte)
        else:
            return ''

    def make_ctrl_network(self):
        network_code = """
        CLR;
        =     "DB_AUSGABE".{0};

        """.format(self.make_operand().symbol)
        return FCNetwork(title=self.name, code=network_code, byte=self.io_out_s.address.byte)

    def __str__(self):
        return str(self.make_network())


class MotorFU(Component):
    FB = 555

    def __init__(self, name, comment, io_out, io_run=None, io_ready=None):
        super(MotorFU, self).__init__(name, comment)
        self.io_out = io_out  # Symbol
        self.io_run = io_run  # Symbol
        self.io_ready = io_ready  # Symbol
        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = SETTINGS['idb_num']+self.io_out.address.get_id()
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            str_io_run = '\n            RM_ANST_EIN              := "{0}",'.format(self.io_run) if self.io_run \
                else '\n            RM_ANST_EIN_FRG          := FALSE,'
            str_io_ready = '\n            FU_BEREIT                := "{0}",'.format(self.io_ready) if self.io_ready \
                else '\n            FU_BEREIT                := TRUE,'

            network_code = """        CALL "MOTOR_1R1G_FU_8P" , "{0}" (
            IMP_ZEITBASIS            := "AUSG_IMP_ZEITBASIS",
            NA_OK                    := "SICHERHEITSKETTE_OK",
            STOER_AUS                := "STOER_AUS",
            QUITT_EX                 := "QUITT_EX",
            ANST_EX_AUTO             := "DB_AUSGABE".{1},{2}
            FRG_TSE                  := TRUE,
            FRG_TSA                  := TRUE,
            MSS                      := FALSE,
            REPS                     := TRUE,
            MANNL                    := TRUE,
            FU_FRG                   := TRUE,{3}
            ZEITBASIS                := "DB_SYSTEM".AUSG_IMP_ZEITBASIS,
            BIBLIOTHEK_ID            := L#{4},
            STOER                    := "DB_AUSG_STOER".{1},
            STOER_SP                 := "DB_AUSG_STOER_SP".{1},
            BA                       := "DB_AUSG_BA".{1},
            AUSGANG                  := "{5}",
            STEHT                    := "DB_AUSG_ENDL_NA".{1},
            RM_EIN                   := "DB_AUSG_ENDL_ANG".{1},
            Para_Werkseinst          := #para_werks);    
// 
      U     #para_werks; 
      O     #para_werks_sammel; 
      SPBN  M{6:02}z;
// 
      CALL "SYS_MOTOR_1R1G_FU_WE" (
           IDB                      := {7},
           Werkseinst               := TRUE,
           SZ_ANST_EIN              := 1.000000e+001,
           SZ_ANST_AUS              := 1.000000e+001,
           SZ_TSE                   := 0.000000e+000,
           SZ_TSA                   := 0.000000e+000,
           GW_LZZ                   := L#0,// 0 = LZZ deaktiviert
           GW_SSZ                   := L#0,// 0 = SSZ deaktiviert
           FU_SZ                    := 1.000000e+001,
           DREH_SZ                  := 0.000000e+000,
           TLS_SZ                   := 0.000000e+000,
           FRG_E_VERZ               := FALSE,
           FRG_A_VERZ               := FALSE,
           STOER_AUS                := FALSE,
           FRG_H_INT                := TRUE,
           FRG_H_EXT                := FALSE,
           ANTR_ST_AUS              := TRUE,
           FUUEB_AKT                := TRUE,
           DRUEB_AKT                := FALSE,
           TLS_AKT                  := FALSE,
           RM_UEB_AKT               := TRUE,
           BED_O_PW                 := FALSE);
// 
M{6:02}z: NOP   0;""".format(self.name, self.make_operand().symbol, str_io_run, str_io_ready, str(self.lib_id),
                             str(self.io_out), self.network_no, self.idb_num)
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)

            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def __str__(self):
        return str(self.make_network())


class Relay(Component):
    FB = None

    def __init__(self, name, comment, io_out, io_on=None):
        super(Relay, self).__init__(name, comment)
        self.io_out = io_out  # Symbol
        self.io_on = io_on  # Symbol
        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = None
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            network_code = """        U     "DB_AUSGABE".{0};
        =     "{1}";""".format(self.make_operand().symbol, str(self.io_out))
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def __str__(self):
        return str(self.make_network())


class Placeholder(Component):
    FB = None

    def __init__(self, name, comment, io_out):
        super(Placeholder, self).__init__(name, comment)
        self.io_out = io_out  # Symbol
        self.lib_id = self.io_out.address.get_id()
        self.network_no = (self.io_out.address.byte % 4) * 8 + self.io_out.address.bit + 2
        self.idb_num = None
        self.fc_num = SETTINGS['visu_fc_num']+self.io_out.address.get_id() // 32
        self.ctrl_fc_num = SETTINGS['ctrl_fc_num']+self.io_out.address.get_id() // 32

    def operands(self):
        return [{"id": self.io_out.address.get_id(), "operand": Operand(self.name, 'BOOL', self.comment)}]

    def make_network(self):
        if self.io_out:
            network_code = """            U     "DB_AUSGABE".{0};
            =     "{1}";""".format(self.make_operand().symbol, str(self.io_out))
            network_title = "A{0}.{1} / {2}".format(str(self.io_out.address.byte), str(self.io_out.address.bit),
                                                    self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def __str__(self):
        return str(self.make_network())


class Measurement:
    FB = 570

    def __init__(self, name, comment, io_in=None, lib_id=0):
        self.name = name
        self.comment = comment
        self.io_in = io_in
        self.lib_id = lib_id
        if io_in:
            self.network_no = io_in.address.get_id() % 16 + 2
            self.idb_num = SETTINGS['idb_num_ai']+io_in.address.get_id()
            self.fc_num = SETTINGS['meas_fc_num']+io_in.address.get_id() // 16

    def to_listbox(self):
        return self.name

    def make_operand(self):
        return Operand(self.name, 'REAL', self.comment)

    def make_network(self):
        if self.io_in:
            network_code = """        CALL "ANALOG_IN_8P" , "{0}" (
           Analogwert               := "{1}",
           Max_Analogwert           := 2.764800e+004,
           IMP_ZEITBASIS            := "AUSG_IMP_ZEITBASIS",
           STOER_AUS                := "STOER_AUS",
           QUITT_EX                 := "QUITT_EX",
           ZEITBASIS                := "DB_SYSTEM".AUSG_IMP_ZEITBASIS,
           BIBLIOTHEK_ID            := L#{2},
           Norm_Wert_WINCC          := "DB_AI_NORM".{3},
           Para_Werkseinst          := #para_werks);
// 
      U     #para_werks; 
      O     #para_werks_sammel; 
      SPBN  M{4:02}z;
// 
      CALL "ANALOG_IN_WE" (
           IDB                      := {5},
           Werkseinst               := TRUE,
           Untergrenze              := 0.000000e+000,// 0,0 l/h
           Obergrenze               := 1.000000e+003,// 1000,0 l/h
           Ersatzwert               := 0.000000e+000,// 0,0 l/h
           Simulation               := 0.000000e+000,// 0,0 l/h
           Offset                   := 0.000000e+000,// 0,0 l/h
           Einheit                  := B#16#E,
           Ber_Unter                := 0.000000e+000,// 0,0 l/h
           Ber_Ueber                := 1.000000e+003,// 1000,0 l/h
           Messber_Unter            := 0.000000e+000,// 0%
           Messber_Ueber            := 1.000000e+002,// 100 %
           MW_Anzahl                := 10,// 10
           MW_Zeit                  := 1.000000e+000,// 1,0 Sekunden
           iorSZ_TSE_DRBR           := 5.000000e+000,// 5 Sekunden
           iorSZ_TSE_UNTER          := 5.000000e+000,// 5 Sekunden
           iorSZ_TSE_UEBER          := 5.000000e+000,// 5 Sekunden
           iodiGW_LZZ               := L#0,// 0 Sekunden = LZZ deaktiviert
           Ueb_DrBr                 := TRUE,
           Ueb_Ber_Unt              := FALSE,
           Ueb_Ber_Ueb              := FALSE,
           Frg_Unterl               := FALSE,
           Frg_Ueberl               := FALSE,
           Frg_MW                   := FALSE);
// 
M{4:02}z: NOP   0;""".format(self.name, str(self.io_in), str(self.lib_id), self.make_operand().symbol, self.network_no,
                             self.idb_num)
            network_title = "PEW{0} / {1}".format(str(self.io_in.address.byte), self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_in.address.byte)
        else:
            return ''

    def make_tiavariables(self):
        variables = """{0}.BIBLIOTHEK_ID	{1}	{2}	{0}.BIBLIOTHEK_ID	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD16	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CHARGE	{1}	{2}	{0}.CHARGE	String	String	18	Binary	Absolute access	%DB{3}.DBX24.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BEREICH	{1}	{2}	{0}.BEREICH	String	String	18	Binary	Absolute access	%DB{3}.DBX44.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.RECHNERNAME	{1}	{2}	{0}.RECHNERNAME	String	String	16	Binary	Absolute access	%DB{3}.DBX64.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CURRENT_USER	{1}	{2}	{0}.CURRENT_USER	String	String	20	Binary	Absolute access	%DB{3}.DBX82.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Norm_Wert_WINCC	{1}	{2}	{0}.Norm_Wert_WINCC	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD104	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Norm_Wert_Eing_WINCC	{1}	{2}	{0}.Norm_Wert_Eing_WINCC	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD108	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwRM_DT	{1}	{2}	{0}.odwRM_DT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD136	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwIN	{1}	{2}	{0}.odwIN	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD140	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwOUT	{1}	{2}	{0}.odwOUT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD144	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwEV_ID	{1}	{2}	{0}.iodwEV_ID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD150	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwDT	{1}	{2}	{0}.iodwDT	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD154	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Untergrenze	{1}	{2}	{0}.Untergrenze	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD158	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Obergrenze	{1}	{2}	{0}.Obergrenze	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD162	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Ersatzwert	{1}	{2}	{0}.Ersatzwert	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD166	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Simulation	{1}	{2}	{0}.Simulation	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD170	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Offset	{1}	{2}	{0}.Offset	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD174	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Einheit	{1}	{2}	{0}.Einheit	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB178	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Ber_Unter	{1}	{2}	{0}.Ber_Unter	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD180	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Ber_Ueber	{1}	{2}	{0}.Ber_Ueber	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD184	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Messber_Unter	{1}	{2}	{0}.Messber_Unter	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD188	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Messber_Ueber	{1}	{2}	{0}.Messber_Ueber	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD192	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.MW_Anzahl	{1}	{2}	{0}.MW_Anzahl	Int	Int	2	Binary	Absolute access	%DB{3}.DBW196	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.MW_Zeit	{1}	{2}	{0}.MW_Zeit	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD198	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE_DRBR	{1}	{2}	{0}.iorSZ_TSE_DRBR	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD202	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE_UNTER	{1}	{2}	{0}.iorSZ_TSE_UNTER	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD206	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE_UEBER	{1}	{2}	{0}.iorSZ_TSE_UEBER	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD210	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiLZZ	{1}	{2}	{0}.iodiLZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD214	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodiGW_LZZ	{1}	{2}	{0}.iodiGW_LZZ	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD218	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owAlarm	{1}	{2}	{0}.owAlarm	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW148	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iowQuitt	{1}	{2}	{0}.iowQuitt	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW222	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False"""\
            .format(self.name, self.__class__.__name__, SETTINGS['tia_connection'], self.idb_num)
        return variables

    def make_tiaalarms(self):
        alarms = '''{0}_Laufzeit	Fehler laufzeit {0}		Errors	"""{0}.owAlarm"""	0	On rising edge	"""{0}.iowQuitt"""	0	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Extern	Fehler extern {0}		Errors	"""{0}.owAlarm"""	1	On rising edge	"""{0}.iowQuitt"""	1	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Wartung	Wartung {0}		Warnings	"""{0}.owAlarm"""	2	On rising edge	"""{0}.iowQuitt"""	2	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Betriebsart	Betriebsart {0}		Hand	"""{0}.owAlarm"""	3	On rising edge	"""{0}.iowQuitt"""	3	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Ansteuerung	Ansteuerung {0}		Ansteurung	"""{0}.owAlarm"""	4	On rising edge	"""{0}.iowQuitt"""	4	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0''' \
            .format(self.name)
        return alarms

    def make_tiaarchive(self):
        archive = '''ANALOG_IN	{0}.Norm_Wert_WINCC	"""{0}.Norm_Wert_WINCC"""	Zyklisch	500 ms	1	500 ms			<Kein>							<Kein>		False	False	True		Aktuell	False	in Prozent (%)	0			System'''\
            .format(self.name)
        return archive

    def __str__(self):
        return str(self.make_network())


class Setpoint:
    FB = 575

    def __init__(self, name, comment, io_out=None, lib_id=0):
        self.name = name
        self.comment = comment
        self.io_out = io_out
        self.lib_id = lib_id
        if io_out:
            self.network_no = io_out.address.get_id() % 16 + 2
            self.idb_num = SETTINGS['idb_num_ao']+io_out.address.get_id()
            self.fc_num = SETTINGS['fc_num_ao']+io_out.address.get_id() // 16

    def to_listbox(self):
        return self.name

    def make_operand(self):
        return Operand(self.name, 'REAL', self.comment)

    def make_network(self):
        if self.io_out:
            network_code = """      L     "DB_AO_NORM".{5}; 
      U     #PROG_TRACK; 
      SPB   M{1:02}a; 
      L     #LMN_REGLER; 
      U     #PROG_REGELUNG; 
      SPB   M{1:02}a; 
      L     0.000000e+000; 
M{1:02}a: NOP   0; 
      T     "DB_AO_NORM".{5}; 
      T     #normwert; 
// 

      CALL "ANALOG_OUT_8P" , "{0}" (
           Normierter_Wert          := #normwert,
           Untergrenze_Eing         := 0.000000e+000,
           Obergrenze_Eing          := 1.000000e+002,
           Max_Analogwert           := 2.764800e+004,
           BIBLIOTHEK_ID            := L#{2},
           Analogwert               := "{3}",
           Para_Werkseinst          := #para_werks);
// 
      U     #para_werks; 
      O     #para_werks_sammel; 
      SPBN  M{1:02}z; 
// 
      CALL "ANALOG_OUT_WE" (
           IDB                      := {4},
           Werkseinst               := TRUE,
           Untergrenze              := 0.000000e+000,
           Obergrenze               := 1.000000e+002,
           Min_Grenze_Ausgang_Proz  := 0.000000e+000,
           Max_Grenze_Ausgang_Proz  := 1.000000e+002,
           Simulation               := 0.000000e+000,
           Offset                   := 0.000000e+000,
           Einheit                  := B#16#1,
           Ber_Unter_Proz           := 0.000000e+000,
           Ber_Ueber_Proz           := 1.000000e+002,
           Glaettung_Sek_Pos_Proz   := 1.000000e+002,
           Glaettung_Sek_Neg_Proz   := 1.000000e+002,
           Frg_Ber_Unter            := FALSE,
           Frg_Ber_Ueber            := FALSE,
           Frg_Glaettung            := FALSE);
// 
M{1:02}z: NOP   0;""".format(self.name, self.network_no, str(self.lib_id), str(self.io_out), self.idb_num, self.make_operand().symbol)
            network_title = "PAW{0} / {1}".format(str(self.io_out.address.byte), self.name)
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def make_tiavariables(self):
            variables = """{0}.BIBLIOTHEK_ID	{1}	{2}	{0}.BIBLIOTHEK_ID	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD18	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BEREICH	{1}	{2}	{0}.BEREICH	String	String	18	Binary	Absolute access	%DB{3}.DBX46.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CHARGE	{1}	{2}	{0}.CHARGE	String	String	18	Binary	Absolute access	%DB{3}.DBX26.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.RECHNERNAME	{1}	{2}	{0}.RECHNERNAME	String	String	16	Binary	Absolute access	%DB{3}.DBX66.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CURRENT_USER	{1}	{2}	{0}.CURRENT_USER	String	String	20	Binary	Absolute access	%DB{3}.DBX84.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Analogwert_Skal	{1}	{2}	{0}.Analogwert_Skal	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD108	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Analogwert_Proz	{1}	{2}	{0}.Analogwert_Proz	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD112	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwRM_DT	{1}	{2}	{0}.odwRM_DT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD122	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwIN	{1}	{2}	{0}.odwIN	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD126	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwOUT	{1}	{2}	{0}.odwOUT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD130	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwEV_ID	{1}	{2}	{0}.iodwEV_ID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD136	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwDT	{1}	{2}	{0}.iodwDT	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD140	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Untergrenze_Ausg	{1}	{2}	{0}.Untergrenze_Ausg	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD144	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Obergrenze_Ausg	{1}	{2}	{0}.Obergrenze_Ausg	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD148	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Min_Grenze_Ausg	{1}	{2}	{0}.Min_Grenze_Ausg	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD152	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Max_Grenze_Ausg	{1}	{2}	{0}.Max_Grenze_Ausg	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD156	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Simulation	{1}	{2}	{0}.Simulation	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD160	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Offset	{1}	{2}	{0}.Offset	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD164	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Einheit	{1}	{2}	{0}.Einheit	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB168	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Ausgabe_Unter	{1}	{2}	{0}.Ausgabe_Unter	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD170	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Ausgabe_Ueber	{1}	{2}	{0}.Ausgabe_Ueber	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD174	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Aend_Ausg_Pos	{1}	{2}	{0}.Aend_Ausg_Pos	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD178	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.Aend_Ausg_Neg	{1}	{2}	{0}.Aend_Ausg_Neg	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD182	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owAlarm	{1}	{2}	{0}.owAlarm	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW134	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iowQuitt	{1}	{2}	{0}.iowQuitt	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW186	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False""" \
                .format(self.name, self.__class__.__name__, SETTINGS['tia_connection'], self.idb_num)
            return variables

    def make_tiaalarms(self):
        alarms = '''{0}_Laufzeit	Fehler laufzeit {0}		Errors	"""{0}.owAlarm"""	0	On rising edge	"""{0}.iowQuitt"""	0	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Extern	Fehler extern {0}		Errors	"""{0}.owAlarm"""	1	On rising edge	"""{0}.iowQuitt"""	1	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Wartung	Wartung {0}		Warnings	"""{0}.owAlarm"""	2	On rising edge	"""{0}.iowQuitt"""	2	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Betriebsart	Betriebsart {0}		Hand	"""{0}.owAlarm"""	3	On rising edge	"""{0}.iowQuitt"""	3	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Ansteuerung	Ansteuerung {0}		Ansteurung	"""{0}.owAlarm"""	4	On rising edge	"""{0}.iowQuitt"""	4	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0''' \
            .format(self.name)
        return alarms

    def make_tiaarchive(self):
        archive = '''ANALOG_OUT	{0}.Analogwert_Skal	"""{0}.Analogwert_Skal"""	Zyklisch	500 ms	1	500 ms			<Kein>							<Kein>		False	False	True		Aktuell	False	in Prozent (%)	0			System'''\
            .format(self.name)
        return archive

    def __str__(self):
        return str(self.make_network())


class AoPlaceholder:
    FB = None

    def __init__(self, name, comment, io_out=None, lib_id=0):
        self.name = name
        self.comment = comment
        self.io_out = io_out
        self.lib_id = lib_id
        if io_out:
            self.network_no = io_out.address.get_id() % 16 + 2
            self.idb_num = SETTINGS['idb_num_ao']+io_out.address.get_id()
            self.fc_num = SETTINGS['fc_num_ao']+io_out.address.get_id() // 16

    def to_listbox(self):
        return self.name

    def make_operand(self):
        return Operand(self.name, 'REAL', self.comment)

    def make_network(self):
        if self.io_out:
            network_code = """      L     0.000000e+000;
      T     "DB_AO_NORM".{2};
      L     0;
      T     "{1}";""".format(self.name, str(self.io_out), self.make_operand().symbol)
            network_title = self.name
            return FCNetwork(title=network_title, code=network_code, byte=self.io_out.address.byte)
        else:
            return ''

    def make_tiavariables(self):
        return ''

    def make_tiaalarms(self):
        return ''

    def make_tiaarchive(self):
        return ''

    def __str__(self):
        return str(self.make_network())


class PID:
    FB = 590
    FB_CTRL = 3500

    def __init__(self, ai, lib_id):
        self.name = "C_"+ai.name
        self.ctrl_idb_name = "IDB_C_"+ai.name
        self.comment = ai.comment
        self.ai_name = ai.name
        self.ai_operand = Operand(ai.name, 'INT', ai.comment).symbol
        self.lib_id = lib_id
        self.idb_num = int(SETTINGS['idb_num_pid'])+lib_id-1
        self.ctrl_idb_num = int(SETTINGS['idb_num_pid_ctrl'])+lib_id-1
        self.fc_num = int(SETTINGS['fc_num_pid'])+lib_id-1

    def to_listbox(self):
        return self.name

    def make_operand(self):
        return Operand(self.name, 'REAL', self.comment)

    def make_fc(self):
        networks = []

        temp_network = FCNetwork(title="Ansteuerung vom Programm ***************************************",
                                 code="""
      U     #PROG_Anst; 
      =     #anst_ex_auto; 
        """)
        networks.append(temp_network)

        temp_network = FCNetwork(title="Normierungsfaktor **********************************************",
                                 code="""
      L     "{0}".iorOBERGR; 
      L     "{0}".iorUNTERGR; 
      -R    ; 
      L     1.000000e+002; // 100 %
      /R    ; 
      T     #norm_faktor; 
        """.format(self.name))
        networks.append(temp_network)

        temp_network = FCNetwork(title="Istwert ********************************************************",
                                 code="""
      L     "DB_AI_NORM".{0}; 
      T     #istwert; 
        """.format(self.ai_operand))
        networks.append(temp_network)

        temp_network = FCNetwork(title="Sollwert vorgeben **********************************************",
                                 code="""
      L     #PROG_EXT_SP; 
      U     #PROG_EXT_SP_Ein; 
      SPB   M04a; 
      L     #PROG_KONST_SP; // Konstant
      U     #PROG_KONST_SP_Ein; 
      SPB   M04a; 
      L     0.000000e+000; 
M04a: T     #sw_a; 
      L     "{0}".HDW.SW_ZS; 
      ==R   ; // Änderungserkennung
      SPB   M04b; 
      L     "{0}".iorOBERGR; 
      L     #sw_a; 
      >=R   ; 
      SPB   M04c; 
      TAK   ; 
M04c: T     "{0}".iorSOLL; 

M04b: NOP   0; 
      L     #sw_a; 
      T     "{0}".HDW.SW_ZS; 
        """.format(self.name))
        networks.append(temp_network)

        temp_network = FCNetwork(title="Handwert vorgeben **********************************************",
                                 code="""
      U     "{0}".HM_REGLER_EIN; // Regler Ein
      SPB   M05a; 
      U     "{0}".HM_B_AUTO; // Regler Betriebsart Automatik
      SPB   M05b; 
      L     "{0}".iorSTELL; 
      SPA   M05c; 
M05b: L     #PROG_FESTWERT1; // Festwertausgabe 1 bei Regler Automatik + Regler aus
      U     #PROG_FEST1_Ein; 
      SPB   M05c; 
      L     #PROG_FESTWERT2; // Festwertausgabe 2 bei Regler Automatik + Regler aus
      U     #PROG_FEST2_Ein; 
      SPB   M05c; 
      L     "{0}".iorSTELL; // Regler einfrieren
      U     #PROG_Freigabe; 
      SPB   M05c; 
      L     0.000000e+000; 
M05c: T     "{1}".MAN; 
M05a: UN    "{0}".HM_B_AUTO; // Regler Betriebsart Automatik
      UN    "{0}".HM_REGLER_EIN; // Regler Ein
      SPB   M05d; 
      L     "{1}".LMN; 
      T     "{0}".iorSTELL; 
M05d: NOP   0; 
        """.format(self.name, self.ctrl_idb_name))
        networks.append(temp_network)

        temp_network = FCNetwork(title="Reglerstörung",
                                 code="""
// Reglerüberschreitung 
      U     "{0}".HM_REGLER_EIN; 
      UN    "{0}".HM_TO_STOER_AUS; 
      =     #frg_reg_stoer_ueber; 
// 
      L     "{0}".iorSOLL; 
      L     "{0}".orREGL_UEB_SW; 
      +R    ; 
      T     #sollwert; 
// 
      CALL "GW_MAX" (
           SOLL                     := #sollwert,
           IST                      := #istwert,
           HYSTERESE                := "{0}".orREGL_UEB_HYST,
           GW_MAX                   := "{0}".HM.GW_REGL_UEB,
           GW_MAX_S                 := #gw_max_s);
// 
      CALL "E_VERZ_I" (
           FRG                      := #frg_reg_stoer_ueber,
           IN                       := "{0}".HM.GW_REGL_UEB,
           TAKT                     := "IMP_1S",
           VZ                       := "{0}".owREGL_UEB_VERZ,
           OUT                      := "{0}".HM.STOER_UEB,
           IZ                       := "{0}".HDW.IZ_STOER_UEB);

// Reglerunterschreitung 
      U     "{0}".HM_REGLER_EIN; 
      UN    "{0}".HM_TO_STOER_AUS; 
      =     #frg_reg_stoer_unter; 
// 
      L     "{0}".iorSOLL; 
      L     "{0}".orREGL_UNT_SW; 
      -R    ; 
      T     #sollwert; 
// 
      CALL "GW_MIN" (
           SOLL                     := #sollwert,
           IST                      := #istwert,
           HYSTERESE                := "{0}".orREGL_UNT_HYST,
           GW_MIN                   := "{0}".HM.GW_REGL_UNT,
           GW_MIN_S                 := #gw_min_s);
// 
      CALL "E_VERZ_I" (
           FRG                      := #frg_reg_stoer_unter,
           IN                       := "{0}".HM.GW_REGL_UNT,
           TAKT                     := "IMP_1S",
           VZ                       := "{0}".owREGL_UNT_VERZ,
           OUT                      := "{0}".HM.STOER_UNT,
           IZ                       := "{0}".HDW.IZ_STOER_UNT);
        """.format(self.name))
        networks.append(temp_network)

        temp_network = FCNetwork(title="Reglerschnittstelle zu WinCC ***********************************",
                                 code="""
      CALL "REGLER_8P" , "{0}" (
           IMP_ZEITBASIS            := "WI_200MS",
           STOER_AUS                := "STOER_AUS",
           QUITT_EX                 := "QUITT_EX",
           ANST_EX_AUTO             := #anst_ex_auto,
           IDB_NR_REGLER            := {1},
           ZEITBASIS                := "DB_SYSTEM".AUSG_IMP_ZEITBASIS,
           FAKT                     := #norm_faktor,
           irISTWERT                := #istwert,
           BIBLIOTHEK_ID            := L#{2},
           Para_Werkseinst          := #para_werks);
// 
      U     "FRG_WERKS_PARA_C"; 
      O     #para_werks; 
      SPBN  M07z; 
// 
      CALL "REGLER_WE" (
           IDB                      := {3},
           Werkseinst               := TRUE,
           SZ_TSE                   := 3.000000e+000,
           SZ_TSA                   := 0.000000e+000,
           OBERGR                   := 1.000000e+002,
           UNTERGR                  := 0.000000e+000,
           P                        := 8.000000e-001,
           I                        := T#15S,
           D                        := T#0MS,
           DV                       := T#0MS,
           TOTZONE                  := 0.000000e+000,
           STELL_BEGR_O             := 1.000000e+002,
           STELL_BEGR_U             := 0.000000e+000,
           STELL_BEGR_ANST          := 1.000000e+001,
           STELL_BEGR_ABST          := 1.000000e+001,
           SW_BEGR_ANST_P           := 1.000000e+001,
           SW_BEGR_ABST_P           := 1.000000e+001,
           SW_BEGR_ANST_N           := 1.000000e+001,
           SW_BEGR_ABST_N           := 1.000000e+001,
           ER_ANZ                   := 5.000000e+000,
           SW_GRENZE_MAX            := 1.000000e+002,
           SW_GRENZE_MIN            := 0.000000e+000,
           REGL_UEB_SW              := 1.000000e+002,
           REGL_UEB_HYST            := 1.000000e+000,
           REGL_UEB_VERZ            := 30,
           REGL_UNT_SW              := 0.000000e+000,
           REGL_UNT_HYST            := 1.000000e+000,
           REGL_UNT_VERZ            := 30,
           EINHEIT_X                := B#16#1,
           EINHEIT_Y                := B#16#1,
           FRG_E_VERZ               := TRUE,
           FRG_A_VERZ               := FALSE,
           STOER_AUS                := TRUE,
           FRG_FRG_LMNROC           := FALSE,
           FRG_FRG_SPROC            := FALSE,
           FRG_DEADB                := FALSE,
           FRG_P                    := TRUE,
           FRG_I                    := TRUE,
           FRG_D                    := FALSE,
           FRG_D_RUECK              := FALSE,
           BED_O_PW                 := FALSE);
// 
M07z: NOP   0; 
        """.format(self.name, self.ctrl_idb_num, self.lib_id, self.idb_num))
        networks.append(temp_network)
        fc_symbol = 'R_{0}'.format(self.name)
        fc_title = 'Regler {0} * {1}'.format(self.ai_name, self.comment)
        fc_name = 'C{}'.format(self.lib_id)
        temp_vars = [
                     "para_werks : BOOL ;	//Werkseinstellung laden",
                     "anst_ex_auto : BOOL ;	//Ansteuerung Extern Automatik",
                     "gw_max_s : BOOL ;	//Grenzwert Max Sicherheitsschaltung",
                    "gw_min_s : BOOL ;	//Grenzwert Min Sicherheitsschaltung",
                    "frg_reg_stoer_ueber : BOOL ;	//Freigabe Reglerstörung Überschreitung",
                    "frg_reg_stoer_unter : BOOL ;	//Freigabe Reglerstörung Unterschreitung",
                    "PROG_Anst : BOOL ;",
                    "PROG_EXT_SP_Ein : BOOL ;",
                    "PROG_KONST_SP_Ein : BOOL ;",
                    "PROG_FEST1_Ein : BOOL ;",
                    "PROG_FEST2_Ein : BOOL ;",
                    "PROG_Freigabe : BOOL ;",
                    "norm_faktor : REAL ;	//Normierungsfaktor",
                    "sw_a : REAL ;	//Sollwert A",
                    "istwert : REAL ;	//Istwert",
                    "sollwert : REAL ;",
                    "PROG_EXT_SP : REAL ;",
                    "PROG_KONST_SP : REAL ;",
                    "PROG_FESTWERT1 : REAL ;",
                    "PROG_FESTWERT2 : REAL ;"]

        return FC(symbol=fc_symbol, num=str(self.fc_num), title=fc_title, author="GEN", family="REGLER", name=fc_name,
                  temp_vars=temp_vars, networks=networks)

    def make_tiavariables(self):
            variables = """{0}.irISTWERT	{1}	{2}	{0}.irISTWERT	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD12	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BIBLIOTHEK_ID	{1}	{2}	{0}.BIBLIOTHEK_ID	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD16	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CHARGE	{1}	{2}	{0}.CHARGE	String	String	18	Binary	Absolute access	%DB{3}.DBX24.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.BEREICH	{1}	{2}	{0}.BEREICH	String	String	18	Binary	Absolute access	%DB{3}.DBX44.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.RECHNERNAME	{1}	{2}	{0}.RECHNERNAME	String	String	16	Binary	Absolute access	%DB{3}.DBX64.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.CURRENT_USER	{1}	{2}	{0}.CURRENT_USER	String	String	20	Binary	Absolute access	%DB{3}.DBX82.0	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.orREGELDIFF	{1}	{2}	{0}.orREGELDIFF	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD112	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.orREGL_UEB_SW	{1}	{2}	{0}.orREGL_UEB_SW	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD116	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.orREGL_UEB_HYST	{1}	{2}	{0}.orREGL_UEB_HYST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD120	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owREGL_UEB_VERZ	{1}	{2}	{0}.owREGL_UEB_VERZ	Int	Int	2	Binary	Absolute access	%DB{3}.DBW124	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.orREGL_UNT_SW	{1}	{2}	{0}.orREGL_UNT_SW	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD126	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.orREGL_UNT_HYST	{1}	{2}	{0}.orREGL_UNT_HYST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD130	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owREGL_UNT_VERZ	{1}	{2}	{0}.owREGL_UNT_VERZ	Int	Int	2	Binary	Absolute access	%DB{3}.DBW134	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwRM_DT	{1}	{2}	{0}.odwRM_DT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD136	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwIN	{1}	{2}	{0}.odwIN	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD140	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.odwOUT	{1}	{2}	{0}.odwOUT	DInt	DInt	4	Binary	Absolute access	%DB{3}.DBD144	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwEV_ID	{1}	{2}	{0}.iodwEV_ID	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD150	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iodwDT	{1}	{2}	{0}.iodwDT	DWord	UDInt	4	Binary	Absolute access	%DB{3}.DBD154	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSE	{1}	{2}	{0}.iorSZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD158	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSE	{1}	{2}	{0}.iorIZ_TSE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD162	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSZ_TSA	{1}	{2}	{0}.iorSZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD166	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorIZ_TSA	{1}	{2}	{0}.iorIZ_TSA	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD170	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSOLL	{1}	{2}	{0}.iorSOLL	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD174	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSTELL	{1}	{2}	{0}.iorSTELL	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD178	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorOBERGR	{1}	{2}	{0}.iorOBERGR	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD182	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorUNTERGR	{1}	{2}	{0}.iorUNTERGR	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD186	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorP	{1}	{2}	{0}.iorP	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD190	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorI	{1}	{2}	{0}.iorI	Time	DInt	4	Binary	Absolute access	%DB{3}.DBD194	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorD	{1}	{2}	{0}.iorD	Time	DInt	4	Binary	Absolute access	%DB{3}.DBD198	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorDV	{1}	{2}	{0}.iorDV	Time	DInt	4	Binary	Absolute access	%DB{3}.DBD202	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorTOTZONE	{1}	{2}	{0}.iorTOTZONE	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD206	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSTELL_BEGR_O	{1}	{2}	{0}.iorSTELL_BEGR_O	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD210	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSTELL_BEGR_U	{1}	{2}	{0}.iorSTELL_BEGR_U	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD214	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSTELL_BEGR_ANST	{1}	{2}	{0}.iorSTELL_BEGR_ANST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD218	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSTELL_BEGR_ABST	{1}	{2}	{0}.iorSTELL_BEGR_ABST	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD222	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_BEGR_ANST_P	{1}	{2}	{0}.iorSW_BEGR_ANST_P	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD226	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_BEGR_ABST_P	{1}	{2}	{0}.iorSW_BEGR_ABST_P	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD230	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_BEGR_ANST_N	{1}	{2}	{0}.iorSW_BEGR_ANST_N	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD234	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_BEGR_ABST_N	{1}	{2}	{0}.iorSW_BEGR_ABST_N	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD238	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorER_ANZ	{1}	{2}	{0}.iorER_ANZ	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD242	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_GRENZE_MAX	{1}	{2}	{0}.iorSW_GRENZE_MAX	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD246	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iorSW_GRENZE_MIN	{1}	{2}	{0}.iorSW_GRENZE_MIN	Real	Real	4	IEEE754	Absolute access	%DB{3}.DBD250	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iobEINHEIT_X	{1}	{2}	{0}.iobEINHEIT_X	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB254	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iobEINHEIT_Y	{1}	{2}	{0}.iobEINHEIT_Y	Byte	USInt	1	Binary	Absolute access	%DB{3}.DBB255	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.owAlarm	{1}	{2}	{0}.owAlarm	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW148	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False
{0}.iowQuitt	{1}	{2}	{0}.iowQuitt	Word	UInt	2	Binary	Absolute access	%DB{3}.DBW256	<No Value>	False	False	<No Value>	<No Value>	Client/Server wide	<No Value>	<No Value>	<No Value>	None	<No Value>	None	<No Value>	False	10	0	100	0	False""" \
                .format(self.name, self.__class__.__name__, SETTINGS['tia_connection'], self.idb_num)
            return variables

    def make_tiaalarms(self):
        alarms = '''{0}_Laufzeit	Fehler laufzeit {0}		Errors	"""{0}.owAlarm"""	0	On rising edge	"""{0}.iowQuitt"""	0	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Extern	Fehler extern {0}		Errors	"""{0}.owAlarm"""	1	On rising edge	"""{0}.iowQuitt"""	1	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Wartung	Wartung {0}		Warnings	"""{0}.owAlarm"""	2	On rising edge	"""{0}.iowQuitt"""	2	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Betriebsart	Betriebsart {0}		Hand	"""{0}.owAlarm"""	3	On rising edge	"""{0}.iowQuitt"""	3	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0
{0}_Ansteuerung	Ansteuerung {0}		Ansteurung	"""{0}.owAlarm"""	4	On rising edge	"""{0}.iowQuitt"""	4	<No value>	0	<No value>	0	False	<No value>	<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>		<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	<No value>	False	0	0	0''' \
            .format(self.name)
        return alarms

    def make_tiaarchive(self):
        archive = '''REGLER	{0}.irISTWERT	"""{0}.irISTWERT"""	Zyklisch	500 ms	1	500 ms			<Kein>							<Kein>		False	False	True		Aktuell	False	in Prozent (%)	0			System			
REGLER	{0}.iorSOLL	"""{0}.iorSOLL"""	Zyklisch	500 ms	1	500 ms			<Kein>							<Kein>		False	False	True		Aktuell	False	in Prozent (%)	0			System			
REGLER	{0}.iorSTELL	"""{0}.iorSTELL"""	Zyklisch	500 ms	1	500 ms			<Kein>							<Kein>		False	False	True		Aktuell	False	in Prozent (%)	0			System'''			\
            .format(self.name)
        return archive

    def __str__(self):
        return str(self.make_fc())


class IDB:
    config = "{ S7_m_c := 'true'; S7_alarm_ui := '0'}"

    def __init__(self, num, fb_num, component_type):
        self.num = num
        self.symbol = 'DB{0}'.format(self.num)
        self.fb_num = fb_num
        self.fb_symbol = 'FB{0}'.format(self.fb_num)
        self.component_type = component_type
        self.header = """
DATA_BLOCK {0}
{1}
{2}
        """.format(self.symbol, self.config, self.fb_symbol)
        self.body = 'BEGIN\nEND_DATA_BLOCK'

    def __str__(self):
        return self.header + self.body

    def to_listbox(self):
        return self.symbol+' - '+self.fb_symbol + ' - ' + self.component_type


class DB:
    def __init__(self, symbol, num='', title='', author='', family='', name='', version='0.1', operands=None,
                 var_type='BOOL', init='FALSE'):
        self.symbol = symbol
        self.num = num
        self.title = title[:80]
        self.author = author
        self.family = family
        self.name = name
        self.version = version
        self.operands = operands
        try:
            for o in self.operands:
                try:
                    o.var_type = var_type
                    o.init = init
                except AttributeError:
                    pass
        except TypeError:
            operands.var_type = var_type
            operands.init = init
        except AttributeError:
            pass
        self.header = """
DATA_BLOCK "{0}"
TITLE = {1}
AUTHOR : {2}
FAMILY : {3}
NAME : {4}
VERSION : {5}
        """.format(self.symbol, self.title, self.author, self.family, self.name, self.version)
        self.definition = ''
        try:
            self.definition = '\n'.join(["\t"+str(o) for o in self.operands])
        except TypeError:
            self.definition = "\t"+str(self.operands)

        self.initiation = ''
        try:
            self.initiation = '\n'.join(["\t"+o.str_init() for o in self.operands])
        except TypeError:
            self.initiation = "\t"+self.operands.str_init()

        self.body = ('\n'*2) \
                    + 'STRUCT\n' + self.definition + '\nEND_STRUCT;' \
                    + '\nBEGIN\n' + self.initiation + '\nEND_DATA_BLOCK'

    def __str__(self):
        return self.header + '\n' + self.body

    def to_listbox(self):
        return 'DB' + str(self.num) + ' - ' + self.symbol


class StoerDB(DB):
    def __init__(self, symbol, num='', title='', author='', family='', name='', version='0.1', operands=None,
                 var_type='BOOL', init='FALSE'):

        neumeldungen = list()

        reserve_bytes = 4
        max_bytes = 300
        reserve = [Operand('r{}_{}'.format(i // 8, i % 8), var_type=var_type, comment='', init=init)
                   for i in range(0, reserve_bytes*8)]
        neumeldungen.append(Struct('SI', 'Sicherungen', reserve))
        neumeldungen.append(Struct('ML', 'Mannloch', reserve))
        neumeldungen.append(Struct('NOT_AUS', 'Not-Aus', reserve))
        neumeldungen.append(Struct('STOER_EXT', 'Externe', reserve))
        neumeldungen.append(Struct('FRG', 'Freigabe/Reparaturschalter', reserve))

        for o in operands:
            o.var_type = var_type
            o.init = init
        neumeldungen.append(Struct('DRAHTBR_ANA', 'Drahtbruch Analogeingaenge', operands))
        neumeldungen.append(Struct('BEREICH_ANA', 'Bereich Analogeingaenge', operands))
        ist_bytes = math.ceil((len(reserve)*5 + len(operands)*2)/8)
        reserve_array = Array('r{}'.format(ist_bytes), ist_bytes, max_bytes-1, 'BYTE', 'B#16#0', comment='Reserve')
        neumeldungen.append(reserve_array)

        altmeldungen = Array('r{}'.format(max_bytes), max_bytes, 2*max_bytes-1, 'BYTE', 'B#16#0', comment='')

        neu = Struct('NEU', 'Neumeldungen', neumeldungen)
        alt = Struct('ALT', 'Altmeldungen', altmeldungen)

        operands = [neu, alt]

        super(StoerDB, self).__init__(symbol, num, title, author, family, name, version, operands, var_type, init)


class FCNetwork:
    def __init__(self, title, code=None, byte=0):
        self.title = title[:80]
        self.code = code
        self.byte = byte
        self.header = """
NETWORK
TITLE ={0}
""".format(self.title)
        if self.code:
            self.body = self.code
        else:
            self.body = """

NOP   0;

            """

    def __str__(self):
        return self.header + '\n' + self.body


class FC:
    def __init__(self, symbol, num='', title='', author='', family='', name='', version='0.1', temp_vars=None, networks=None):
        self.symbol = symbol
        self.num = num
        self.title = title[:80]
        self.author = author
        self.family = family
        self.name = name
        self.version = version
        self.temp_vars = temp_vars
        self.networks = networks

        self.header = """
FUNCTION "{0}" : VOID
TITLE = {1}

AUTHOR : {2}
FAMILY : {3}
NAME : {4}
VERSION : {5}
        """.format(self.symbol, self.title, self.author, self.family, self.name, self.version)
        if temp_vars:
            self.vars = '\nVAR_TEMP\n' + '\n'.join(["\t"+str(v) for v in self.temp_vars]) + '\nEND_VAR'
        else:
            self.vars = ''
        if networks:
            self.body = '\nBEGIN\n' + '\n'.join(["\t"+str(n) for n in self.networks]) + '\nEND_FUNCTION'
        else:
            self.body = '\nBEGIN\n' + str(FCNetwork('Netzwerk 0')) + '\nEND_FUNCTION'

    def __str__(self):
        return self.header + '\n' + self.vars + '\n' + self.body

    def to_listbox(self):
        return "FC{0}".format(self.num) + ' - ' + self.symbol


def xls_get_symbols(path):
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    symbols = [Symbol(ws.cell(r, 1).value, ws.cell(r, 4).value, ws.cell(r, 3).value, ws.cell(r, 2).value)
               for r in range(1, ws.max_row)]
    symbols.sort(key=lambda x: (x.address.area, x.address.byte, x.address.bit if x.address.bit is not None else 0))
    return symbols


def txt_get_symbols(path):
    with open(path, 'r') as textFile:
        lines = textFile.readlines()

        symbols = [Symbol(l.split('\t')[0], l.split('\t')[3][:-2], l.split('\t')[2], l.split('\t')[1])
                   for l in lines if len(l.split('\t')) == 4]
        symbols.sort(key=lambda x: (x.address.area, x.address.byte, x.address.bit if x.address.bit is not None else 0))
        return symbols


def get_outputs(symbols):
    outputs = []
    for s in symbols:
        io_add = IOAddress(s[1])
        if io_add.area == 'A' or io_add.area == 'Q':
            outputs.append(s)
    return outputs


def operands_from_objects(do=None, ai=None, ao=None):
    if ai:
        byte_range = range(SETTINGS['start_analog'], SETTINGS['end_analog']+1, 2)
        operands = [Operand(str(IOAddress('PEW', i)), 'REAL', '', '0.0') for i in byte_range]
        for obj in ai:
            if obj.io_in.is_analog_input() and (obj.lib_id < len(operands)):
                operands[obj.lib_id] = obj.make_operand()
        return operands
    elif ao:
        byte_range = range(SETTINGS['start_analog'], SETTINGS['end_analog']+1, 2)
        operands = [Operand(str(IOAddress('PAW', i)), 'REAL', '', '0.0') for i in byte_range]
        for obj in ao:
            if obj.io_out.is_analog_output() and (obj.lib_id < len(operands)):
                operands[obj.lib_id] = obj.make_operand()
        return operands
    elif do:
        byte_range_do1 = range(SETTINGS['start_digital1']*8, SETTINGS['start_analog']*8)
        operands = [Operand(str(IOAddress('A', i // 8, i % 8)), 'BOOL', '', 'FALSE') for i in byte_range_do1]
        byte_range_do2 = range(SETTINGS['start_digital2']*8, (SETTINGS['end_digital2']+1)*8)
        operands += [Operand(str(IOAddress('A', i // 8, i % 8)), 'BOOL', '', 'FALSE') for i in byte_range_do2]
        for obj in do:
            for o in obj.operands():
                if o['id'] < len(operands):
                    operands[o['id']] = o['operand']
        return operands


def lib_from_symbols(symbols):

    byte_range_do1 = range(SETTINGS['start_digital1']*8, SETTINGS['start_analog']*8)
    output_symbols = [Symbol(str(IOAddress('A', i // 8, i % 8)), '', 'BOOL', 'A', i // 8, i % 8)
                      for i in byte_range_do1]
    byte_range_do2 = range(SETTINGS['start_digital2']*8, (SETTINGS['end_digital2']+1)*8)
    output_symbols += [Symbol(str(IOAddress('A', i // 8, i % 8)), '', 'BOOL', 'A', i // 8, i % 8)
                       for i in byte_range_do2]
    byte_range_ai = range(SETTINGS['start_analog'], SETTINGS['end_analog']+1, 2)
    ai_symbols = [Symbol(str(IOAddress('PEW', i)), '', 'INT', 'PEW', i) for i in byte_range_ai]
    byte_range_ao = range(SETTINGS['start_analog'], SETTINGS['end_analog']+1, 2)
    ao_symbols = [Symbol(str(IOAddress('PAW', i)), '', 'INT', 'PAW', i) for i in byte_range_ao]

    for s in symbols:
        if s.is_output() and (s.address.get_id() < len(output_symbols)):
            output_symbols[s.address.get_id()] = s
        elif s.is_analog_input() and (s.address.get_id() < len(ai_symbols)):
            ai_symbols[s.address.get_id()] = s
        elif s.is_analog_output() and (s.address.get_id() < len(ao_symbols)):
            ao_symbols[s.address.get_id()] = s

    return [s.to_lib_address() for s in output_symbols], [s.to_lib_comment() for s in output_symbols], \
           [s.to_lib_address() for s in ai_symbols], [s.to_lib_comment() for s in ai_symbols],\
           [s.to_lib_address() for s in ao_symbols], [s.to_lib_comment() for s in ao_symbols]


def lib_from_controllers(controllers):
    return ["C{0:03}".format(c.lib_id) for c in controllers], [c.comment for c in controllers]


def operands_from_components(components):
    return [c.make_operand() for c in components]


def components_from_symbols(symbols, byte_range=None):

    components = []
    component_symbols = [s for s in symbols if s.is_component() and (s.address.byte in byte_range if byte_range else True)
               and (s.address.byte in range(SETTINGS['start_digital1'], SETTINGS['start_analog'])
                    or s.address.byte in range(SETTINGS['start_digital2'], SETTINGS['end_digital2']))]
    for s in component_symbols:
        s.make_component(components=components, names=[c.name for c in components], symbols=symbols)

    # update components by searching inputs
    input_symbols = [s for s in symbols if s.is_input()]
    for s in input_symbols:
        s.update_component(components=components, names=[c.name for c in components])

    return components


def measurements_from_symbols(symbols):
    return [s.make_measurement() for s in symbols if s.is_analog_input() and
            s.address.byte in range(SETTINGS['start_analog'], SETTINGS['end_analog']+1)]


def pid_from_measurements(measurements):
    lib_id = 1
    pid_list = []
    for m in measurements:
        if m.io_in.is_pid():
            pid_list.append(PID(m, lib_id))
            lib_id += 1
    return pid_list


def analog_outputs_from_symbols(symbols):
    return [s.make_analog_output() for s in symbols if s.is_analog_output()
            and s.address.byte >= SETTINGS['start_analog']]


def visu_networks_from_objects(objects):
    networks = {}
    for o in objects:
        fc_num = o.fc_num
        if fc_num not in networks:
            networks[fc_num] = []
        networks[fc_num].append(o.make_network())
    return networks


def ctrl_networks_from_components(components):
    networks = {}
    for c in components:
        fc_num = c.ctrl_fc_num
        if fc_num not in networks:
            networks[fc_num] = []
        networks[fc_num].append(c.make_ctrl_network())
    return networks


def visu_fcs_from_components(components):
    networks = visu_networks_from_objects(components)
    init_network = FCNetwork(title="Werkseinstellung laden !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!",
                             code="""
        U     "FRG_WERKS_PARA_DO";
        =     #para_werks_sammel; 
        """)
    fcs = []
    symbols = []
    for fc in list(networks.keys()):
        networks[fc].insert(0, init_network)
        fc_num = int(fc)
        first_byte = networks[fc][1].byte // 4 * 4
        fc_symbol = 'A{0}-A{1}'.format(first_byte, first_byte+3)
        fc_title = 'Ausgabe Byte A{0} - A{1}'.format(first_byte, first_byte+3)
        fc_name = 'A{}'.format(first_byte)
        temp_vars = ["para_werks : BOOL ;	//Werkseinstellung laden selektiv",
                     "para_werks_sammel : BOOL ;	//Werkseinstellung laden alle"]

        symbols.append(Symbol(fc_symbol, fc_title, None, 'FC', fc_num))
        fcs.append(FC(symbol=fc_symbol, num=fc_num, title=fc_title, author="GEN", family="OUT", name=fc_name,
                      temp_vars=temp_vars, networks=networks[fc]))
    return fcs, symbols


def ctrl_fcs_from_components(components):
    networks = ctrl_networks_from_components(components)
    fcs = []
    symbols = []
    for fc in list(networks.keys()):
        fc_num = int(fc)
        # try:
        first_byte = networks[fc][0].byte // 4 * 4
        # except IndexError:
        #    print(" ctrl fcs from components Index Error fc: {0}, networks: {1}, this network {2}".format(fc, networks, networks[fc]))

        fc_symbol = 'AM_A{0}-A{1}'.format(first_byte, first_byte+3)
        fc_title = 'Ausgabematrix Byte A{0} - A{1}'.format(first_byte, first_byte+3)
        fc_name = 'A{}'.format(first_byte)
        symbols.append(Symbol(fc_symbol, fc_title, None, 'FC', fc_num))
        fcs.append(FC(symbol=fc_symbol, num=fc_num, title=fc_title, author="GEN", family="OUT", name=fc_name,
                      networks=networks[fc]))
    return fcs, symbols


def tiavariables_from_objects(objects):
    variables = []
    symbols = []
    for c in objects:
        if c.make_tiavariables() != "":
            name = c.lib_id
            symbols.append(Symbol(c.name,"", None, 'TIA {0}'.format(c.lib_id)))
            variables.append(c.make_tiavariables())
    return variables, symbols

def tiaalarms_from_objects(objects):
    alarms = []
    symbols = []
    for c in objects:
        if c.make_tiavariables() != "":
            name = c.lib_id
            symbols.append(Symbol(c.name,"", None, 'TIA_AL {0}'.format(c.lib_id)))
            alarms.append(c.make_tiaalarms())
    return alarms, symbols

def tiaarchive_from_objects(objects):
    archives = []
    symbols = []
    for o in objects:
        archive = o.make_tiaarchive()
        if archive != "":
            name = o.lib_id
            symbols.append(Symbol(o.name,"", None, 'TIA_AR {0}'.format(o.lib_id)))
            archives.append(archive)
    return archives, symbols

def meas_fcs_from_measurements(measurements):
    networks = visu_networks_from_objects(measurements)
    init_network = FCNetwork(title="Werkseinstellung laden !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!",
                             code="""
        U     "FRG_WERKS_PARA_AI";
        =     #para_werks_sammel; 
        """)
    fcs = []
    symbols = []
    for fc in list(networks.keys()):
        networks[fc].insert(0, init_network)
        fc_num = int(fc)
        first_byte = networks[fc][1].byte // 16 * 16
        fc_symbol = 'AI_NORM_PEW{0}_{1}'.format(first_byte, first_byte+30)
        fc_title = 'Analog Eingabe  PEW{0}-{1}'.format(first_byte, first_byte+30)
        fc_name = 'PEW{}'.format(first_byte)
        temp_vars = ["para_werks : BOOL ;	//Werkseinstellung laden selektiv",
                     "para_werks_sammel : BOOL ;	//Werkseinstellung laden alle"]

        symbols.append(Symbol(fc_symbol, fc_title, None, 'FC', fc_num))
        fcs.append(FC(symbol=fc_symbol, num=str(fc_num), title=fc_title, author="GEN", family="AI", name=fc_name,
                      temp_vars=temp_vars, networks=networks[fc]))
    return fcs, symbols


def fcs_from_pid_list(pid_list):
    fcs = []
    symbols = []
    for p in pid_list:
        fc = p.make_fc()
        symbols.append(Symbol(fc.symbol, fc.title, None, 'FC', fc.num))
        fcs.append(fc)
    return fcs, symbols


def ao_fcs_from_ao(analog_outputs):
    networks = visu_networks_from_objects(analog_outputs)
    init_network = FCNetwork(title="Werkseinstellung laden !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!",
                             code="""
        U     "FRG_WERKS_PARA_AO";
        =     #para_werks_sammel; 
        """)

    fcs = []
    symbols = []
    for fc in list(networks.keys()):
        networks[fc].insert(0, init_network)
        fc_num = int(fc)
        first_byte = networks[fc][1].byte // 16 * 16
        fc_symbol = 'AO_NORM_PAW{0}_{1}'.format(first_byte, first_byte+30)
        fc_title = 'Analog Ausgabe  PAW{0}-{1}'.format(first_byte, first_byte+30)
        fc_name = 'PAW{}'.format(first_byte)
        temp_vars = ["normwert : REAL ;",
                     "para_werks : BOOL ;	//Werkseinstellung laden selektiv",
                     "para_werks_sammel : BOOL ;	//Werkseinstellung laden alle",
                     "PROG_TRACK : BOOL ;",
                     "LMN_REGLER : REAL ;",
                     "PROG_REGELUNG : BOOL ;"]

        symbols.append(Symbol(fc_symbol, fc_title, None, 'FC', fc_num))
        fcs.append(FC(symbol=fc_symbol, num=str(fc_num), title=fc_title, author="GEN", family="AO", name=fc_name,
                      temp_vars=temp_vars, networks=networks[fc]))
    return fcs, symbols

# def meas_fcs_from_measurements(measurements):
#     networks = meas_networks_from_measurements(measurements)
#     fcs = []
#     symbols = []
#     fc_num = int(SETTINGS['meas_fc_num'])
#     fc_symbol = 'IN_ANALOG'
#     fc_title = 'Analogwerte einlesen'
#     fc_name = 'A_IN'
#     symbols.append(Symbol(fc_symbol, fc_title, None, 'FC', fc_num))
#     fcs.append(FC(symbol=fc_symbol, num=fc_num, title=fc_title, author="GEN", family="IN", name=fc_name,
#                   networks=networks))
#     return fcs, symbols


def idbs_from_objects(objects):
    idbs = []
    symbols = []
    for o in objects:
        try:
            if o.FB:
                idb_symbol = Symbol(o.name, o.comment, 'FB{}'.format(o.FB), 'DB', o.idb_num)
                symbols.append(idb_symbol)
                idbs.append(IDB(o.idb_num, o.FB, o.__class__.__name__))
        except AttributeError:
            pass
        try:
            if o.FB_CTRL:
                idb_symbol = Symbol(o.ctrl_idb_name, o.comment, 'FB{}'.format(o.FB_CTRL), 'DB', o.ctrl_idb_num)
                symbols.append(idb_symbol)
                idbs.append(IDB(o.ctrl_idb_num, o.FB_CTRL, 'CTRL_{}'.format(o.__class__.__name__)))
        except AttributeError:
            pass
    return idbs, symbols


def dbs_from_objects(db_config, do=None, ai=None, ao=None):
    operands = operands_from_objects(do, ai, ao)
    dbs = []
    symbols = []
    for db in db_config:
        db_name, db_num, db_comment, db_family, db_name_short, var_type, init = db
        dbs.append(DB(symbol=db_name, num=db_num, title=db_comment, author="GEN", family=db_family, name=db_name_short,
                      operands=operands, var_type=var_type, init=init))
        symbols.append(Symbol(db_name, db_comment, 'DB{}'.format(db_num), 'DB', db_num))
    return dbs, symbols


def strdbs_from_symbols(db_config, do=None, ai=None, ao=None):
    operands = operands_from_objects(do, ai, ao)
    dbs = []
    symbols = []
    for db in db_config:
        db_name, db_num, db_comment, var_type, init = db
        dbs.append(StoerDB(symbol=db_name, num=db_num, title=db_comment, author="GEN", family="OUT", name="DIV",
                           operands=operands, var_type=var_type, init=init))
        symbols.append(Symbol(db_name, db_comment, 'DB{}'.format(db_num), 'DB', db_num))
    return dbs, symbols


if __name__ == '__main__':
    print('Nothing to do here, just a class container')

